import pandas as pd
from flask import Flask, render_template, flash, url_for, request, make_response, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
import os, time
import io
import base64
import json
import datetime
import xlsxwriter
import os, sys, glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side
# from flask import send_from_directory
from flask import send_file
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from fnmatch import fnmatch
import time
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


REGULAR_SIZE = 11
REGULAR_FONT = 'Cambria'


class Sheet_Generation:
    def __init__(self):
        self = self

    @staticmethod
    def create_equipment_sheet(wb,ws,temp_df):
    
        new_header      = temp_df.iloc[0] 
        temp_df         = temp_df[1:] 
        temp_df.columns = new_header
        temp_df.dropna(axis=1, how='all',inplace=True)           
        ws.title = 'Equipment List '    
        file_name = "Cleaning_room_report_{}.xlsx".format(str(datetime.datetime.today().strftime('%d_%m_%Y')))    
        end_column =temp_df.shape[1] +2   
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        ws["B" + str(2)] = 'ANNEXURE - I'
        ws["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = ws["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        ws["B" + str(3)] = 'Equipment List '
        ws["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = ws["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')


        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=end_column)
        ws["B" + str(5)] = 'List of Equipments and their Product Contact Surface Area  '
        ws["B" + str(5)].fill = PatternFill(start_color='ff99ff', end_color='ff99ff', fill_type="solid")
        currentCell = ws["B" + str(5)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = 6
        j=1
        for row_data in temp_df.columns:
            ws.cell(row=row, column=j+2, value=row_data)
            ws.cell(row=row, column=j+2).fill = PatternFill(start_color='66ccff', end_color='66ccff', fill_type="solid")
            j=j+1
        row=row+1
        ws['B6']="SR.no"
        ws['B6'].fill = PatternFill(start_color='66ccff', end_color='66ccff', fill_type="solid")
        for row_data in temp_df.itertuples():
            for j in range(0,len(row_data) ):   
                
                if j ==0:
                    ws.cell(row=row, column=j+2, value=row_data[j])
                    ws.cell(row=row, column=j+2).fill = PatternFill(start_color='669999', end_color='669999', fill_type="solid")
                if j==1:
                    ws.cell(row=row, column=j+2, value=row_data[j])
                    ws.cell(row=row, column=j+2).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
                   
                if j>1:
                    ws.cell(row=row, column=j+2, value=float(row_data[j]))
                    ws.cell(row=row, column=j+2).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
            row=row+1    

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws["B" + str(row)] = 'Total surface Area '
        ws["B" + str(row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        column = 4
        while column < end_column+1:
            i = get_column_letter(column)
            ws['{}{}'.format(i,row)] = "=SUM({}7:{}{})".format(i,i,row-1)
            ws['{}{}'.format(i,row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
            column += 1  
        
        column = 4
        sheet_ranges = wb.active
        sheet_ranges.column_dimensions["B"].width = 10
        sheet_ranges.column_dimensions["C"].width = 35    
        while column < end_column:
            i = get_column_letter(column)
            ws.column_dimensions[i].width = 10
            column += 1   
            
        return wb
        
    @staticmethod    
    def create_product_sheet(wb,ws,product_frame):
        product_sheet = wb.create_sheet('Product Details')  
        
        end_column =product_frame.shape[1] +2   
        product_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        product_sheet["B" + str(2)] = 'ANNEXURE - II'
        product_sheet["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = product_sheet["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        product_sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        product_sheet["B" + str(3)] = 'Product Details '
        product_sheet["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = product_sheet["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')    
        row = 5
        j=1    
         
        for row_data in product_frame.columns:
            product_sheet.cell(row=row, column=j+2, value=row_data)
            product_sheet.cell(row=row, column=j+2).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            j=j+1
        row=row+1
        product_sheet['B5']="SR.no"
        product_sheet['B5'].fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
        for row_data in product_frame.itertuples():
            for j in range(0,len(row_data) ):   
                product_sheet.cell(row=row, column=j+2, value=row_data[j])
                if j ==0:
                    product_sheet.cell(row=row, column=j+2).fill = PatternFill(start_color='669999', end_color='669999', fill_type="solid")
                else :
                    product_sheet.cell(row=row, column=j+2).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            row=row+1 

        sheet_ranges = wb['Product Details']
        sheet_ranges.column_dimensions["C"].width = 20
        sheet_ranges.column_dimensions["D"].width = 20
        sheet_ranges.column_dimensions["E"].width = 15
        sheet_ranges.column_dimensions["F"].width = 20
        sheet_ranges.column_dimensions["G"].width = 20
        sheet_ranges.column_dimensions["I"].width = 20
        sheet_ranges.column_dimensions["J"].width = 20
        sheet_ranges.column_dimensions["K"].width = 20
        sheet_ranges.column_dimensions["L"].width = 20

        return wb
       
    @staticmethod
    def create_pde_sheet(wb,ws,input_frame,product_frame):
        pde_sheet  = wb.create_sheet('PDE')
        end_column = len(product_frame.PDE_VALUE)+4
        pde_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        pde_sheet["B" + str(2)] = 'ANNEXURE - III'
        pde_sheet["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = pde_sheet["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')


        pde_sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        pde_sheet["B" + str(3)] = 'MACO Calculation on based PDE Value  '
        pde_sheet["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = pde_sheet["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')    
        row = 5
        j=1

        product_list  = input_frame.iloc[0][2:]
        for product in product_list:
            pde_sheet.cell(row=row, column=j+4, value=product)
            pde_sheet.cell(row=row, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            noel_value = product_frame[product_frame['Product_Name']==product]['NOEL'].values[0]
            pde_sheet.cell(row=row+1, column=j+4, value=noel_value)
            pde_sheet.cell(row=row+1, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            j=j+1


        pde_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        pde_sheet["B" + str(5)] = 'Product Name (A) '
        pde_sheet["B" + str(5)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = pde_sheet["B" + str(5)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')          
        
        row=row+1
        
        pde_sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
        pde_sheet["B" + str(6)] = 'PDE Value (A) '
        pde_sheet["B" + str(6)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = pde_sheet["B" + str(6)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')  

        row=row+1
        pde_sheet.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=2)
        pde_sheet["B" + str(row)] = 'S.r. No'
        pde_sheet["B" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = pde_sheet["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        pde_sheet.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
        pde_sheet["C" + str(row)] = 'Product Name'
        pde_sheet["C" + str(row)].fill = PatternFill(start_color='cc7a00', end_color='cc7a00', fill_type="solid")
        currentCell = pde_sheet["C" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        pde_sheet.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=4)
        pde_sheet["D" + str(row)] = 'Generic Name'
        pde_sheet["D" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = pde_sheet["D" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        row=row+2
        
        counter  = 0
        
        new_header      = input_frame.iloc[0] 
        temp_df         = input_frame[1:] 
        temp_df.columns = new_header
        temp_df.dropna(axis=1, how='all',inplace=True) 

        temp_row_count =0
        for product in product_list:
        
            pde_sheet.cell(row=row, column=2, value=counter)
            pde_sheet.cell(row=row, column=2).fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
            
            pde_sheet.cell(row=row, column=3, value=product)
            pde_sheet.cell(row=row, column=3).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
            pde_sheet.cell(row=row, column=4, value=product)
            pde_sheet.cell(row=row, column=4).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
    
            #print("**********************************************************************")
            #print(input_frame.shape[0])
            sumation_row =input_frame.shape[0]+6
            temp_column_count = 0
            for temp_product in product_list:
                i = get_column_letter(temp_column_count+5)
                pde_sheet['{}{}'.format(i,row)] = """=({column}{row}* 'Product Details'!G{product_list_row} *1000)/('Product Details'!K{product_list_row}* 'Equipment List '!{column_name}{sumation_row} *100)""".format(column=i,column_name=get_column_letter(temp_column_count+4),
                                                                                                                        row=6,
                                                                                                                        sumation_row=sumation_row,
                                                                                                                        product_list_row=temp_row_count+6)
                temp_column_count =temp_column_count+1
                
            row=row+1
            temp_row_count = temp_row_count+1
            
          
            
            
            
        pde_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        pde_sheet["B" + str(row)] = ' Minimum Value MACO Based on PDE value  '
        pde_sheet["B" + str(row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        for column in range(5, len(product_list)+5):
            i = get_column_letter(column)
            pde_sheet['{}{}'.format(i,row)] = "=MIN({}7:{}{})".format(i,i,row-1)
            pde_sheet['{}{}'.format(i,row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        
        sheet_ranges = wb['PDE']
        sheet_ranges.column_dimensions["B"].width = 10
        sheet_ranges.column_dimensions["C"].width = 20
        sheet_ranges.column_dimensions["D"].width = 20
        column = 5
        end_column = column+len(product_list)
        while column < end_column:
            i = get_column_letter(column)
            sheet_ranges.column_dimensions[i].width = 10
            column += 1   
        
        return wb

    @staticmethod
    def create_toxicity_sheet(wb,ws,input_frame,product_frame):
        toxicity_sheet  = wb.create_sheet('Toxicity')
        end_column = len(product_frame.NOEL)+4
        toxicity_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        toxicity_sheet["B" + str(2)] = 'ANNEXURE - IV'
        toxicity_sheet["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = toxicity_sheet["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')


        toxicity_sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        toxicity_sheet["B" + str(3)] = 'MACO Calculation on based Toxicity '
        toxicity_sheet["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = toxicity_sheet["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')    
        row = 5
        j=1
        
        
        
        product_list  = input_frame.iloc[0][2:]
        for product in product_list:
            toxicity_sheet.cell(row=row, column=j+4, value=product)
            toxicity_sheet.cell(row=row, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            noel_value = product_frame[product_frame['Product_Name']==product]['NOEL'].values[0]
            toxicity_sheet.cell(row=row+1, column=j+4, value=noel_value)
            toxicity_sheet.cell(row=row+1, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            j=j+1


        toxicity_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        toxicity_sheet["B" + str(5)] = 'Product Name (A) '
        toxicity_sheet["B" + str(5)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = toxicity_sheet["B" + str(5)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')          
        
        row=row+1
        
        toxicity_sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
        toxicity_sheet["B" + str(6)] = 'NOEL (A) '
        toxicity_sheet["B" + str(6)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = toxicity_sheet["B" + str(6)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 




        row=row+1
        toxicity_sheet.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=2)
        toxicity_sheet["B" + str(row)] = 'S.r. No'
        toxicity_sheet["B" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = toxicity_sheet["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        toxicity_sheet.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
        toxicity_sheet["C" + str(row)] = 'Product Name'
        toxicity_sheet["C" + str(row)].fill = PatternFill(start_color='cc7a00', end_color='cc7a00', fill_type="solid")
        currentCell = toxicity_sheet["C" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        toxicity_sheet.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=4)
        toxicity_sheet["D" + str(row)] = 'Generic Name'
        toxicity_sheet["D" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = toxicity_sheet["D" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        row=row+2




        counter  = 0        
        new_header      = input_frame.iloc[0] 
        temp_df         = input_frame[1:] 
        temp_df.columns = new_header
        temp_df.dropna(axis=1, how='all',inplace=True) 

        temp_row_count =0
        for product in product_list:
        
            toxicity_sheet.cell(row=row, column=2, value=counter)
            toxicity_sheet.cell(row=row, column=2).fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
            
            toxicity_sheet.cell(row=row, column=3, value=product)
            toxicity_sheet.cell(row=row, column=3).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
            toxicity_sheet.cell(row=row, column=4, value=product)
            toxicity_sheet.cell(row=row, column=4).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
    
            #print("**********************************************************************")
            #print(input_frame.shape[0])
            sumation_row =input_frame.shape[0]+6
            temp_column_count = 0
            for temp_product in product_list:
                i = get_column_letter(temp_column_count+5)
                toxicity_sheet['{}{}'.format(i,row)] = """=({column}{row}* 'Product Details'!G{product_list_row} *1000)/('Product Details'!K{product_list_row}* 'Equipment List '!{column_name}{sumation_row} *100)""".format(column=i,column_name=get_column_letter(temp_column_count+4),
                                                                                                                        row=6,
                                                                                                                        sumation_row=sumation_row,
                                                                                                                        product_list_row=temp_row_count+6)
                temp_column_count =temp_column_count+1
                
            row=row+1
            temp_row_count = temp_row_count+1
            
            
        toxicity_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        toxicity_sheet["B" + str(row)] = ' Minimum Value MACO Based on toxicity '
        toxicity_sheet["B" + str(row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        for column in range(5, len(product_list)+5):
            i = get_column_letter(column)
            toxicity_sheet['{}{}'.format(i,row)] = "=MIN({}7:{}{})".format(i,i,row-1)
            toxicity_sheet['{}{}'.format(i,row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        
        sheet_ranges = wb['Toxicity']
        sheet_ranges.column_dimensions["B"].width = 10
        sheet_ranges.column_dimensions["C"].width = 30
        sheet_ranges.column_dimensions["D"].width = 30

        column = 5
        end_column=column +len(product_list)
        while column < end_column:
            i = get_column_letter(column)
            sheet_ranges.column_dimensions[i].width = 30
            column += 1   
        
        return wb

    @staticmethod
    def create_dose_base_sheet(wb,ws,input_frame,product_frame):
        dose_base_sheet  = wb.create_sheet('dose base')
        end_column = len(product_frame.MRDD)+4
        dose_base_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        dose_base_sheet["B" + str(2)] = 'ANNEXURE - V'
        dose_base_sheet["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = dose_base_sheet["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')


        dose_base_sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        dose_base_sheet["B" + str(3)] = 'MACO Calculation on Dose Base'
        dose_base_sheet["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = dose_base_sheet["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')    
        row = 5
        j=1


        product_list  = input_frame.iloc[0][2:]
        for product in product_list:
            dose_base_sheet.cell(row=row, column=j+4, value=product)
            dose_base_sheet.cell(row=row, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            MRDD_value = product_frame[product_frame['Product_Name']==product]['MRDD'].values[0]
            dose_base_sheet.cell(row=row+1, column=j+4, value=MRDD_value)
            dose_base_sheet.cell(row=row+1, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            j=j+1


        dose_base_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        dose_base_sheet["B" + str(5)] = 'Product Name (A) '
        dose_base_sheet["B" + str(5)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = dose_base_sheet["B" + str(5)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')          
        
        row=row+1
        
        dose_base_sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
        dose_base_sheet["B" + str(6)] = 'MRDD (A) '
        dose_base_sheet["B" + str(6)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = dose_base_sheet["B" + str(6)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        
        
        
        row=row+1
        dose_base_sheet.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=2)
        dose_base_sheet["B" + str(row)] = 'S.r. No'
        dose_base_sheet["B" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = dose_base_sheet["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        dose_base_sheet.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
        dose_base_sheet["C" + str(row)] = 'Product Name'
        dose_base_sheet["C" + str(row)].fill = PatternFill(start_color='cc7a00', end_color='cc7a00', fill_type="solid")
        currentCell = dose_base_sheet["C" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        dose_base_sheet.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=4)
        dose_base_sheet["D" + str(row)] = 'Generic Name'
        dose_base_sheet["D" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = dose_base_sheet["D" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        row=row+2
        counter  = 0        
        new_header      = input_frame.iloc[0] 
        temp_df         = input_frame[1:] 
        temp_df.columns = new_header
        temp_df.dropna(axis=1, how='all',inplace=True) 

        temp_row_count =0
        for product in product_list:
        
            dose_base_sheet.cell(row=row, column=2, value=counter)
            dose_base_sheet.cell(row=row, column=2).fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
            
            dose_base_sheet.cell(row=row, column=3, value=product)
            dose_base_sheet.cell(row=row, column=3).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
            dose_base_sheet.cell(row=row, column=4, value=product)
            dose_base_sheet.cell(row=row, column=4).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
    
 
            sumation_row =input_frame.shape[0]+6
            temp_column_count = 0
            for temp_product in product_list:
                i = get_column_letter(temp_column_count+5)
                dose_base_sheet['{}{}'.format(i,row)] = """=({column}{row}* 'Product Details'!H{product_list_row} *1000)/('Product Details'!J{product_list_row}* 'Equipment List '!{column_name}{sumation_row} *100)""".format(column=i,column_name=get_column_letter(temp_column_count+4),
                                                                                                                        row=6,
                                                                                                                        sumation_row=sumation_row,
                                                                                                                        product_list_row=temp_row_count+6)
                temp_column_count =temp_column_count+1
                
            row=row+1
            temp_row_count = temp_row_count+1
            
            
            
        dose_base_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        dose_base_sheet["B" + str(row)] = 'Minimum Value MACO Based on dose base'
        dose_base_sheet["B" + str(row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        for column in range(5, len(product_list)+5):
            i = get_column_letter(column)
            dose_base_sheet['{}{}'.format(i,row)] = "=MIN({}7:{}{})".format(i,i,row-1)
            dose_base_sheet['{}{}'.format(i,row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        
        sheet_ranges = wb['dose base']
        sheet_ranges.column_dimensions["B"].width = 10
        sheet_ranges.column_dimensions["C"].width = 30
        sheet_ranges.column_dimensions["D"].width = 30

        column = 5
        end_column=column +len(product_list)
        while column < end_column:
            i = get_column_letter(column)
            sheet_ranges.column_dimensions[i].width = 30
            column += 1   
        
        return wb

    @staticmethod
    def create_ppm_sheet(wb,ws,input_frame,product_frame):
        ppm_sheet  = wb.create_sheet('10 ppm')
        end_column = len(product_frame.MRDD)+4
        ppm_sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
        ppm_sheet["B" + str(2)] = 'ANNEXURE - V'
        ppm_sheet["B" + str(2)].fill = PatternFill(start_color='00cc99', end_color='00cc99', fill_type="solid")
        currentCell = ppm_sheet["B" + str(2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')


        ppm_sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=end_column)
        ppm_sheet["B" + str(3)] = 'Amt. of Residue allowed (G/CM)  Calculation on based 10 PPM'
        ppm_sheet["B" + str(3)].fill = PatternFill(start_color='ff9933', end_color='ff9933', fill_type="solid")
        currentCell = ppm_sheet["B" + str(3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')    
        row = 5
        j=1


        product_list  = input_frame.iloc[0][2:]
        for product in product_list:
            ppm_sheet.cell(row=row, column=j+4, value=product)
            ppm_sheet.cell(row=row, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            ppm_sheet.cell(row=row+1, column=j+4, value=0.001)
            ppm_sheet.cell(row=row+1, column=j+4).fill = PatternFill(start_color='ffcc99', end_color='ffcc99', fill_type="solid")
            j=j+1


        ppm_sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        ppm_sheet["B" + str(5)] = 'Product Name (A) '
        ppm_sheet["B" + str(5)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = ppm_sheet["B" + str(5)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')          
        
        row=row+1
        
        ppm_sheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
        ppm_sheet["B" + str(6)] = '10 ppm  (A) '
        ppm_sheet["B" + str(6)].fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
        currentCell = ppm_sheet["B" + str(6)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        
        
        
        row=row+1
        ppm_sheet.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=2)
        ppm_sheet["B" + str(row)] = 'S.r. No'
        ppm_sheet["B" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = ppm_sheet["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        ppm_sheet.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
        ppm_sheet["C" + str(row)] = 'Product Name'
        ppm_sheet["C" + str(row)].fill = PatternFill(start_color='cc7a00', end_color='cc7a00', fill_type="solid")
        currentCell = ppm_sheet["C" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 

        ppm_sheet.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=4)
        ppm_sheet["D" + str(row)] = 'Generic Name'
        ppm_sheet["D" + str(row)].fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
        currentCell = ppm_sheet["D" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        row=row+2
        counter  = 0        
        new_header      = input_frame.iloc[0] 
        temp_df         = input_frame[1:] 
        temp_df.columns = new_header
        temp_df.dropna(axis=1, how='all',inplace=True) 

        temp_row_count =0
        for product in product_list:
        
            ppm_sheet.cell(row=row, column=2, value=counter)
            ppm_sheet.cell(row=row, column=2).fill = PatternFill(start_color='e0e0d1', end_color='e0e0d1', fill_type="solid")
            
            ppm_sheet.cell(row=row, column=3, value=product)
            ppm_sheet.cell(row=row, column=3).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
            ppm_sheet.cell(row=row, column=4, value=product)
            ppm_sheet.cell(row=row, column=4).fill = PatternFill(start_color='3399ff', end_color='3399ff', fill_type="solid")
            
    
            #print("**********************************************************************")
            #print(input_frame.shape[0])
            sumation_row =input_frame.shape[0]+6
            temp_column_count = 0
            for temp_product in product_list:
                i = get_column_letter(temp_column_count+5)
                ppm_sheet['{}{}'.format(i,row)] = """=({column}{row}* 'Product Details'!H{product_list_row} *1000)/( 'Equipment List '!{column_name}{sumation_row})""".format(column=i,column_name=get_column_letter(temp_column_count+4),
                                                                                                                        row=6,
                                                                                                                        sumation_row=sumation_row,
                                                                                                                        product_list_row=temp_row_count+6)
                temp_column_count =temp_column_count+1
                
            row=row+1
            temp_row_count = temp_row_count+1
            
            
            
        ppm_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ppm_sheet["B" + str(row)] = ' Minimum Value Amt. of Residue allowed (G/CM)  Based on 10pmm '
        ppm_sheet["B" + str(row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center') 
        for column in range(5, len(product_list)+5):
            i = get_column_letter(column)
            ppm_sheet['{}{}'.format(i,row)] = "=MIN({}7:{}{})".format(i,i,row-1)
            ppm_sheet['{}{}'.format(i,row)].fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type="solid")
        
        sheet_ranges = wb['10 ppm']
        sheet_ranges.column_dimensions["B"].width = 10
        sheet_ranges.column_dimensions["C"].width = 30
        sheet_ranges.column_dimensions["D"].width = 30

        column = 5
        end_column=column +len(product_list)
        while column < end_column:
            i = get_column_letter(column)
            sheet_ranges.column_dimensions[i].width = 30
            column += 1   
        
        return wb
